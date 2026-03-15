"""
ETF Exporter Module

Exports ETF analysis results to Excel and CSV formats with:
- Color-coded scores
- Formatted cells
- Multiple sheets for detailed data
- Summary statistics

Classes:
    SpreadsheetExporter: Export analysis results to Excel/CSV

Example:
    >>> from src.exporter import SpreadsheetExporter
    >>> exporter = SpreadsheetExporter()
    >>> filepath = exporter.export_to_excel(etfs_data, strategies_data)
"""

import os
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Color,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class SpreadsheetExporter:
    """
    Exports ETF analysis results to Excel and CSV formats.

    Features:
        - Color-coded cells based on scores
        - Formatted headers and data
        - Multiple sheets (Summary, Details, Statistics)
        - Excel tables with auto-filter

    Example:
        >>> exporter = SpreadsheetExporter()
        >>> filepath = exporter.export_to_excel(etfs_data, strategies_data)
        >>> print(f"Exported to: {filepath}")
    """

    # Color definitions
    COLORS = {
        "high": "00C853",  # Green for high scores (8-10)
        "medium": "FFB300",  # Amber for medium scores (5-7)
        "low": "F44336",  # Red for low scores (0-4)
        "header": "1E3A8A",  # Dark blue for headers
        "header_text": "FFFFFF",  # White text for headers
        "alternating1": "F5F5F5",  # Light gray for alternating rows
        "alternating2": "FFFFFF",  # White for alternating rows
    }

    def __init__(self, output_dir: str = "output"):
        """
        Initialize the exporter.

        Args:
            output_dir (str): Directory to save exported files
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _get_score_color(self, score: float) -> str:
        """Get color hex code based on score value."""
        if score >= 8:
            return self.COLORS["high"]
        elif score >= 5:
            return self.COLORS["medium"]
        else:
            return self.COLORS["low"]

    def _apply_cell_style(self, cell, score: Optional[float] = None, is_header: bool = False):
        """Apply styling to a cell based on its content."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if is_header:
            cell.font = Font(bold=True, color=self.COLORS["header_text"])
            cell.fill = PatternFill(start_color=self.COLORS["header"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        elif score is not None:
            cell.fill = PatternFill(
                start_color=self._get_score_color(score), fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF" if score >= 5 else "000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def export_to_excel(
        self, etfs_data: list, strategies_data: list, filename: Optional[str] = None
    ) -> str:
        """
        Export ETF analysis results to Excel format.

        Args:
            etfs_data (list): List of ETF data dictionaries
            strategies_data (list): List of strategy analysis dictionaries
            filename (Optional[str]): Custom filename (default: auto-generated)

        Returns:
            str: Path to the exported Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ETF_Analysis_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        # Create workbook
        wb = Workbook()

        # Create Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, etfs_data, strategies_data)

        # Create Details Sheet
        ws_details = wb.create_sheet("Details")
        self._create_details_sheet(ws_details, etfs_data, strategies_data)

        # Create Statistics Sheet
        ws_stats = wb.create_sheet("Statistics")
        self._create_statistics_sheet(ws_stats, strategies_data)

        # Create Strategy Scores Sheet
        ws_scores = wb.create_sheet("Strategy Scores")
        self._create_scores_sheet(ws_scores, etfs_data, strategies_data)

        # Save workbook
        wb.save(filepath)
        return filepath

    def _create_summary_sheet(self, ws, etfs_data: list, strategies_data: list):
        """Create the summary sheet with top ETFs."""
        # Headers
        headers = ["Rank", "Ticker", "Name", "Category", "Price", "Total Score", "Avg Score", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_cell_style(cell, is_header=True)

        # Prepare data sorted by total score
        data = []
        for etf, strategies in zip(etfs_data, strategies_data):
            data.append({
                "ticker": etf.get("symbol", ""),
                "name": etf.get("name", "")[:40],
                "category": etf.get("category", ""),
                "price": etf.get("price"),
                "total_score": strategies.get("total_score", 0),
                "avg_score": strategies.get("average_score", 0),
            })

        data.sort(key=lambda x: x["total_score"], reverse=True)

        # Add data rows
        for row_idx, row_data in enumerate(data[:50], 2):  # Top 50
            rank = row_idx - 1
            recommendation = self._get_recommendation(row_data["total_score"])

            row_values = [
                rank,
                row_data["ticker"],
                row_data["name"],
                row_data["category"],
                f"${row_data['price']:.2f}" if row_data["price"] else "N/A",
                row_data["total_score"],
                f"{row_data['avg_score']:.1f}",
                recommendation,
            ]

            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if col == 6:  # Total Score column
                    self._apply_cell_style(cell, score=row_data["total_score"] / 10)
                elif col == 8:  # Recommendation column
                    rec_score = row_data["total_score"]
                    if rec_score >= 70:
                        self._apply_cell_style(cell, score=8)
                    elif rec_score >= 50:
                        self._apply_cell_style(cell, score=6)
                    else:
                        self._apply_cell_style(cell, score=3)

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions[get_column_letter(3)].width = 40  # Name column

    def _create_details_sheet(self, ws, etfs_data: list, strategies_data: list):
        """Create the details sheet with all ETF metrics."""
        # Headers
        headers = [
            "Ticker", "Name", "Category", "Family", "Price", "NAV", "Expense Ratio",
            "Dividend Yield", "YTD Return", "3Y Return", "5Y Return", "Beta",
            "52W High", "52W Low", "50D Avg", "200D Avg", "Holdings", "AUM",
            "Volume", "Avg Volume", "Total Score", "Avg Score"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_cell_style(cell, is_header=True)

        # Add data
        for row_idx, (etf, strategies) in enumerate(zip(etfs_data, strategies_data), 2):
            row_values = [
                etf.get("symbol", ""),
                etf.get("name", "")[:50],
                etf.get("category", ""),
                etf.get("family", ""),
                etf.get("price"),
                etf.get("nav_price"),
                etf.get("expense_ratio"),
                etf.get("dividend_yield"),
                etf.get("ytd_return"),
                etf.get("three_year_return"),
                etf.get("five_year_return"),
                etf.get("beta"),
                etf.get("52_week_high"),
                etf.get("52_week_low"),
                etf.get("50_day_avg"),
                etf.get("200_day_avg"),
                etf.get("holdings_count"),
                etf.get("market_cap"),
                etf.get("volume"),
                etf.get("avg_volume"),
                strategies.get("total_score", 0),
                strategies.get("average_score", 0),
            ]

            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if col in [21, 22]:  # Score columns
                    score = value / 10 if value else 0
                    self._apply_cell_style(cell, score=score)

            # Alternating row colors
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.fill.start_color.rgb == "00000000":
                        cell.fill = PatternFill(
                            start_color=self.COLORS["alternating1"], fill_type="solid"
                        )

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions[get_column_letter(2)].width = 35

    def _create_statistics_sheet(self, ws, strategies_data: list):
        """Create the statistics sheet with score distributions."""
        from src.strategies import STRATEGY_NAMES

        ws.cell(row=1, column=1, value="ETF Analysis Statistics")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Summary stats
        total_etfs = len(strategies_data)
        scores = [s.get("total_score", 0) for s in strategies_data]
        avg_score = sum(scores) / len(scores) if scores else 0
        max_score = max(scores) if scores else 0
        min_score = min(scores) if scores else 0

        stats = [
            ("Total ETFs Analyzed", total_etfs),
            ("Average Total Score", f"{avg_score:.1f}"),
            ("Highest Score", max_score),
            ("Lowest Score", min_score),
            ("Strong Buys (>=70)", sum(1 for s in scores if s >= 70)),
            ("Moderate Buys (50-69)", sum(1 for s in scores if 50 <= s < 70)),
            ("Hold/Weak (<50)", sum(1 for s in scores if s < 50)),
        ]

        for row, (label, value) in enumerate(stats, 3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)

        # Strategy averages
        ws.cell(row=12, column=1, value="Strategy Average Scores")
        ws.cell(row=12, column=1).font = Font(bold=True, size=12)

        strategy_scores = {}
        for strategies in strategies_data:
            for strategy_name in STRATEGY_NAMES.keys():
                if strategy_name not in strategy_scores:
                    strategy_scores[strategy_name] = []
                score = strategies.get(strategy_name, {}).get("score", 0)
                strategy_scores[strategy_name].append(score)

        for row, (strategy_name, score_list) in enumerate(strategy_scores.items(), 14):
            avg = sum(score_list) / len(score_list) if score_list else 0
            display_name = STRATEGY_NAMES.get(strategy_name, strategy_name)
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=2, value=f"{avg:.2f}")
            ws.cell(row=row, column=3, value=f"{avg/10:.1%}")

            # Color code
            score_cell = ws.cell(row=row, column=3)
            self._apply_cell_style(score_cell, score=avg/10)

    def _create_scores_sheet(self, ws, etfs_data: list, strategies_data: list):
        """Create the strategy scores sheet."""
        from src.strategies import STRATEGY_NAMES

        # Headers
        headers = ["Ticker"] + list(STRATEGY_NAMES.values()) + ["Total", "Average"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_cell_style(cell, is_header=True)

        # Add data
        for row_idx, (etf, strategies) in enumerate(zip(etfs_data, strategies_data), 2):
            row_values = [etf.get("symbol", "")]

            for strategy_name in STRATEGY_NAMES.keys():
                score = strategies.get(strategy_name, {}).get("score", 0)
                row_values.append(score)

            row_values.extend([
                strategies.get("total_score", 0),
                f"{strategies.get('average_score', 0):.1f}"
            ])

            for col, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if col > 1 and col <= len(STRATEGY_NAMES) + 1:  # Strategy score columns
                    self._apply_cell_style(cell, score=value/10)
                elif col > len(STRATEGY_NAMES) + 1:  # Total/Average columns
                    if isinstance(value, (int, float)):
                        score = value / 10 if value <= 10 else value / 100
                        self._apply_cell_style(cell, score=score)

        # Add table
        table = Table(
            displayName="StrategyScores",
            ref=f"A1:{get_column_letter(len(headers))}{len(etfs_data) + 1}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    def _get_recommendation(self, score: float) -> str:
        """Get recommendation string based on score."""
        if score >= 70:
            return "Strong Buy"
        elif score >= 50:
            return "Buy"
        elif score >= 30:
            return "Hold"
        else:
            return "Avoid"

    def export_to_csv(
        self, etfs_data: list, strategies_data: list, filename: Optional[str] = None
    ) -> str:
        """
        Export ETF analysis results to CSV format.

        Args:
            etfs_data (list): List of ETF data dictionaries
            strategies_data (list): List of strategy analysis dictionaries
            filename (Optional[str]): Custom filename

        Returns:
            str: Path to the exported CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ETF_Analysis_{timestamp}.csv"

        filepath = os.path.join(self.output_dir, filename)

        # Create DataFrame
        from src.strategies import STRATEGY_NAMES

        data = []
        for etf, strategies in zip(etfs_data, strategies_data):
            row = {
                "Ticker": etf.get("symbol", ""),
                "Name": etf.get("name", ""),
                "Category": etf.get("category", ""),
                "Family": etf.get("family", ""),
                "Price": etf.get("price"),
                "Expense_Ratio": etf.get("expense_ratio"),
                "Dividend_Yield": etf.get("dividend_yield"),
                "YTD_Return": etf.get("ytd_return"),
                "Three_Year_Return": etf.get("three_year_return"),
                "Five_Year_Return": etf.get("five_year_return"),
                "Beta": etf.get("beta"),
                "Holdings_Count": etf.get("holdings_count"),
                "AUM": etf.get("market_cap"),
                "Total_Score": strategies.get("total_score", 0),
                "Average_Score": strategies.get("average_score", 0),
                "Recommendation": self._get_recommendation(strategies.get("total_score", 0)),
            }

            # Add individual strategy scores
            for strategy_name in STRATEGY_NAMES.keys():
                score = strategies.get(strategy_name, {}).get("score", 0)
                row[strategy_name] = score

            data.append(row)

        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False)
        return filepath
