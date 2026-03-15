"""
Mutual Fund Backtester Module

Backtests mutual fund investment strategies on historical data with:
- Configurable date ranges
- Rebalance frequencies
- Position sizing
- Transaction costs
- Benchmark comparison

Classes:
    BacktestConfig: Configuration for backtest parameters
    BacktestResults: Results from backtest execution
    Backtester: Main backtesting engine

Example:
    >>> from src.backtester import Backtester, BacktestConfig
    >>> config = BacktestConfig(start_date="2023-01-01", end_date="2024-01-01")
    >>> backtester = Backtester(config)
    >>> results = backtester.run_backtest(fund_tickers)
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd

from src.data_fetcher import MutualFundDataFetcher
from src.strategies import MutualFundStrategies


@dataclass
class BacktestConfig:
    """Configuration for backtest parameters."""

    start_date: str
    end_date: str
    initial_capital: float = 100000
    rebalance_frequency: str = "monthly"  # daily, weekly, monthly, quarterly, yearly
    position_size_pct: float = 0.10  # Max 10% per position
    min_score_threshold: float = 50  # Minimum score to buy
    transaction_cost_pct: float = 0.001  # 0.1% per trade
    benchmark: str = "VOO"  # Benchmark fund/ETF


@dataclass
class Trade:
    """Represents a single trade."""

    date: datetime
    ticker: str
    action: str  # "BUY" or "SELL"
    shares: int
    price: float
    value: float
    transaction_cost: float
    reason: str


@dataclass
class BacktestResults:
    """Results from backtest execution."""

    config: BacktestConfig
    total_return: float = 0.0
    annualized_return: float = 0.0
    benchmark_return: float = 0.0
    sharpe_ratio: float = 0.0
    sortino_ratio: float = 0.0
    max_drawdown: float = 0.0
    volatility: float = 0.0
    beta: float = 0.0
    alpha: float = 0.0
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    total_transaction_costs: float = 0.0
    portfolio_values: pd.DataFrame = field(default_factory=pd.DataFrame)
    trades: list = field(default_factory=list)

    @property
    def excess_return(self) -> float:
        """Return excess return over benchmark."""
        return self.total_return - self.benchmark_return


class Backtester:
    """
    Backtests ETF investment strategies on historical data.

    Features:
        - Configurable rebalance frequencies
        - Position sizing constraints
        - Transaction cost modeling
        - Benchmark comparison
        - Performance metrics calculation

    Example:
        >>> config = BacktestConfig(
        ...     start_date="2023-01-01",
        ...     end_date="2024-01-01",
        ...     initial_capital=100000
        ... )
        >>> backtester = Backtester(config)
        >>> results = backtester.run_backtest(["SPY", "QQQ", "VTI"])
    """

    def __init__(self, config: BacktestConfig):
        """
        Initialize the backtester.

        Args:
            config (BacktestConfig): Backtest configuration
        """
        self.config = config
        self.fetcher = MutualFundDataFetcher()
        self.trades: list[Trade] = []
        self.portfolio_values: list[dict] = []

    def _get_rebalance_dates(self) -> list[datetime]:
        """Generate rebalance dates based on frequency."""
        start = datetime.strptime(self.config.start_date, "%Y-%m-%d")
        end = datetime.strptime(self.config.end_date, "%Y-%m-%d")

        dates = []
        current = start

        while current <= end:
            dates.append(current)

            if self.config.rebalance_frequency == "daily":
                current += timedelta(days=1)
            elif self.config.rebalance_frequency == "weekly":
                current += timedelta(weeks=1)
            elif self.config.rebalance_frequency == "monthly":
                # Add 1 month
                if current.month == 12:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=current.month + 1)
            elif self.config.rebalance_frequency == "quarterly":
                # Add 3 months
                new_month = current.month + 3
                new_year = current.year + (new_month - 1) // 12
                new_month = ((new_month - 1) % 12) + 1
                current = current.replace(year=new_year, month=new_month)
            elif self.config.rebalance_frequency == "yearly":
                current = current.replace(year=current.year + 1)

        return dates

    def _get_historical_price(self, ticker: str, date: datetime) -> Optional[float]:
        """Get historical price for a mutual fund on a specific date."""
        try:
            # Fetch with a date range around the target date
            yf_ticker = yf.Ticker(ticker)
            hist = yf_ticker.history(
                start=(date - timedelta(days=5)).strftime("%Y-%m-%d"),
                end=(date + timedelta(days=1)).strftime("%Y-%m-%d"),
            )
            if hist is not None and len(hist) > 0:
                # Get closest date
                return float(hist["Close"].iloc[-1])
        except Exception:
            pass
        return None

    def _calculate_strategy_scores(self, ticker: str, date: datetime) -> dict:
        """Calculate strategy scores for a mutual fund at a point in time."""
        # For simplicity, use current data
        # In production, you'd fetch historical fundamentals
        fund_data = self.fetcher.fetch_fund_data(ticker)
        if fund_data:
            return MutualFundStrategies.analyze_fund(fund_data)
        return {}

    def _rebalance_portfolio(
        self,
        date: datetime,
        current_holdings: dict,
        cash: float,
        portfolio_value: float,
        prices: dict,
    ) -> tuple[dict, float, float]:
        """
        Rebalance portfolio based on strategy scores.

        Returns:
            tuple: (new_holdings, new_cash, transaction_costs)
        """
        new_holdings = {}
        transaction_costs = 0.0

        # Get scores for all ETFs
        scored_etfs = []
        for ticker in prices.keys():
            scores = self._calculate_strategy_scores(ticker, date)
            total_score = scores.get("total_score", 0)
            if total_score >= self.config.min_score_threshold:
                scored_etfs.append((ticker, total_score))

        # Sort by score
        scored_etfs.sort(key=lambda x: x[1], reverse=True)

        # Select top ETFs (diversify)
        max_positions = int(1 / self.config.position_size_pct)
        selected = scored_etfs[:max_positions]

        if not selected:
            # Sell all holdings if no good opportunities
            for ticker, shares in current_holdings.items():
                if shares > 0 and ticker in prices:
                    price = prices[ticker]
                    sell_value = shares * price
                    cost = sell_value * self.config.transaction_cost_pct
                    transaction_costs += cost
                    self.trades.append(Trade(
                        date=date,
                        ticker=ticker,
                        action="SELL",
                        shares=shares,
                        price=price,
                        value=sell_value,
                        transaction_cost=cost,
                        reason="Below threshold"
                    ))
            return new_holdings, cash + sum(
                shares * prices.get(ticker, 0)
                for ticker, shares in current_holdings.items()
            ) - transaction_costs, transaction_costs

        # Calculate equal weight for each selected ETF
        target_value = portfolio_value * self.config.position_size_pct

        # Sell holdings not in selected
        for ticker, shares in current_holdings.items():
            if shares > 0 and ticker not in [t for t, _ in selected]:
                if ticker in prices:
                    price = prices[ticker]
                    sell_value = shares * price
                    cost = sell_value * self.config.transaction_cost_pct
                    transaction_costs += cost
                    cash += sell_value - cost
                    self.trades.append(Trade(
                        date=date,
                        ticker=ticker,
                        action="SELL",
                        shares=shares,
                        price=price,
                        value=sell_value,
                        transaction_cost=cost,
                        reason="Not in top picks"
                    ))

        # Buy/update selected positions
        for ticker, score in selected:
            if ticker in prices:
                price = prices[ticker]
                target_shares = int(target_value / price)

                current_shares = current_holdings.get(ticker, 0)
                shares_diff = target_shares - current_shares

                if shares_diff > 0:
                    # Buy
                    buy_value = shares_diff * price
                    cost = buy_value * self.config.transaction_cost_pct
                    if cash >= buy_value + cost:
                        cash -= buy_value + cost
                        transaction_costs += cost
                        new_holdings[ticker] = target_shares
                        self.trades.append(Trade(
                            date=date,
                            ticker=ticker,
                            action="BUY",
                            shares=shares_diff,
                            price=price,
                            value=buy_value,
                            transaction_cost=cost,
                            reason=f"Score: {score}"
                        ))
                elif shares_diff < 0:
                    # Sell
                    sell_value = abs(shares_diff) * price
                    cost = sell_value * self.config.transaction_cost_pct
                    cash += sell_value - cost
                    transaction_costs += cost
                    new_holdings[ticker] = target_shares
                    self.trades.append(Trade(
                        date=date,
                        ticker=ticker,
                        action="SELL",
                        shares=abs(shares_diff),
                        price=price,
                        value=sell_value,
                        transaction_cost=cost,
                        reason="Rebalance"
                    ))
                else:
                    # Hold
                    new_holdings[ticker] = current_shares

        return new_holdings, cash, transaction_costs

    def run_backtest(self, tickers: list) -> BacktestResults:
        """
        Run the backtest on a list of ETF tickers.

        Args:
            tickers (list): List of ETF ticker symbols

        Returns:
            BacktestResults: Backtest results with performance metrics
        """
        import yfinance as yf

        # Initialize
        cash = self.config.initial_capital
        holdings = {ticker: 0 for ticker in tickers}
        rebalance_dates = self._get_rebalance_dates()

        # Fetch benchmark data
        benchmark = yf.Ticker(self.config.benchmark)
        benchmark_hist = benchmark.history(
            start=self.config.start_date,
            end=self.config.end_date,
        )
        benchmark_start_price = benchmark_hist["Close"].iloc[0] if len(benchmark_hist) > 0 else 1

        # Fetch all ETF historical data
        etf_data = {}
        for ticker in tickers:
            try:
                etf = yf.Ticker(ticker)
                hist = etf.history(
                    start=self.config.start_date,
                    end=self.config.end_date,
                )
                etf_data[ticker] = hist
            except Exception:
                pass

        # Track daily portfolio values
        all_dates = pd.date_range(
            start=self.config.start_date,
            end=self.config.end_date,
            freq="D"
        )

        for current_date in all_dates:
            # Get current prices
            prices = {}
            for ticker in tickers:
                if ticker in etf_data and len(etf_data[ticker]) > 0:
                    hist = etf_data[ticker]
                    mask = hist.index <= current_date
                    if mask.sum() > 0:
                        prices[ticker] = float(hist[mask]["Close"].iloc[-1])

            # Calculate portfolio value
            portfolio_value = cash + sum(
                shares * prices.get(ticker, 0)
                for ticker, shares in holdings.items()
            )

            # Record portfolio value
            self.portfolio_values.append({
                "date": current_date,
                "portfolio_value": portfolio_value,
                "cash": cash,
                "holdings_value": portfolio_value - cash,
            })

            # Rebalance if needed
            if current_date in rebalance_dates:
                holdings, cash, _ = self._rebalance_portfolio(
                    current_date, holdings, cash, portfolio_value, prices
                )

        # Calculate final metrics
        return self._calculate_results(benchmark_hist, benchmark_start_price)

    def _calculate_results(
        self, benchmark_hist: pd.DataFrame, benchmark_start_price: float
    ) -> BacktestResults:
        """Calculate final backtest metrics."""
        if not self.portfolio_values:
            return BacktestResults(config=self.config)

        # Convert to DataFrame
        pv_df = pd.DataFrame(self.portfolio_values)
        pv_df.set_index("date", inplace=True)

        # Portfolio returns
        initial_value = self.config.initial_capital
        final_value = pv_df["portfolio_value"].iloc[-1]
        total_return = (final_value - initial_value) / initial_value

        # Annualized return
        days = (pv_df.index[-1] - pv_df.index[0]).days
        annualized_return = (final_value / initial_value) ** (365 / days) - 1 if days > 0 else 0

        # Daily returns
        daily_returns = pv_df["portfolio_value"].pct_change().dropna()

        # Volatility (annualized)
        volatility = daily_returns.std() * (252 ** 0.5) if len(daily_returns) > 0 else 0

        # Sharpe Ratio (assuming 5% risk-free rate)
        risk_free_rate = 0.05
        excess_return = annualized_return - risk_free_rate
        sharpe_ratio = excess_return / volatility if volatility > 0 else 0

        # Sortino Ratio (downside deviation)
        downside_returns = daily_returns[daily_returns < 0]
        downside_deviation = downside_returns.std() * (252 ** 0.5) if len(downside_returns) > 0 else 0
        sortino_ratio = excess_return / downside_deviation if downside_deviation > 0 else 0

        # Maximum Drawdown
        rolling_max = pv_df["portfolio_value"].cummax()
        drawdown = (pv_df["portfolio_value"] - rolling_max) / rolling_max
        max_drawdown = drawdown.min()

        # Benchmark return
        benchmark_end_price = benchmark_hist["Close"].iloc[-1] if len(benchmark_hist) > 0 else 1
        benchmark_return = (benchmark_end_price - benchmark_start_price) / benchmark_start_price

        # Beta and Alpha
        if len(daily_returns) > 10:
            # Get benchmark daily returns
            benchmark_returns = benchmark_hist["Close"].pct_change().dropna()

            # Align dates
            common_dates = daily_returns.index.intersection(benchmark_returns.index)
            if len(common_dates) > 10:
                port_ret = daily_returns.loc[common_dates]
                bench_ret = benchmark_returns.loc[common_dates]

                # Calculate beta
                covariance = port_ret.cov(bench_ret)
                benchmark_variance = bench_ret.var()
                beta = covariance / benchmark_variance if benchmark_variance > 0 else 0

                # Calculate alpha (annualized)
                alpha = annualized_return - (risk_free_rate + beta * (benchmark_return - risk_free_rate))
            else:
                beta = 0
                alpha = 0
        else:
            beta = 0
            alpha = 0

        # Trade statistics
        total_trades = len(self.trades)
        buy_trades = [t for t in self.trades if t.action == "BUY"]
        sell_trades = [t for t in self.trades if t.action == "SELL"]

        # Simple win/loss calculation based on profitable sells
        winning_trades = sum(1 for t in sell_trades if t.transaction_cost < t.value * 0.01)
        losing_trades = len(sell_trades) - winning_trades
        win_rate = winning_trades / len(sell_trades) if sell_trades else 0

        # Total transaction costs
        total_transaction_costs = sum(t.transaction_cost for t in self.trades)

        return BacktestResults(
            config=self.config,
            total_return=total_return,
            annualized_return=annualized_return,
            benchmark_return=benchmark_return,
            sharpe_ratio=sharpe_ratio,
            sortino_ratio=sortino_ratio,
            max_drawdown=max_drawdown,
            volatility=volatility,
            beta=beta,
            alpha=alpha,
            total_trades=total_trades,
            winning_trades=winning_trades,
            losing_trades=losing_trades,
            win_rate=win_rate,
            total_transaction_costs=total_transaction_costs,
            portfolio_values=pv_df,
            trades=self.trades,
        )

    def export_results(self, results: BacktestResults, filename: Optional[str] = None) -> str:
        """Export backtest results to Excel."""
        from openpyxl import Workbook

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ETF_Backtest_{timestamp}.xlsx"

        filepath = os.path.join("output", filename)
        os.makedirs("output", exist_ok=True)

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_data = [
            ("Backtest Period", f"{self.config.start_date} to {self.config.end_date}"),
            ("Initial Capital", f"${self.config.initial_capital:,.2f}"),
            ("Final Value", f"${self.config.initial_capital * (1 + results.total_return):,.2f}"),
            ("Total Return", f"{results.total_return * 100:.1f}%"),
            ("Annualized Return", f"{results.annualized_return * 100:.1f}%"),
            ("Benchmark Return", f"{results.benchmark_return * 100:.1f}%"),
            ("Excess Return", f"{results.excess_return * 100:.1f}%"),
            ("", ""),
            ("Sharpe Ratio", f"{results.sharpe_ratio:.2f}"),
            ("Sortino Ratio", f"{results.sortino_ratio:.2f}"),
            ("Max Drawdown", f"{results.max_drawdown * 100:.1f}%"),
            ("Volatility", f"{results.volatility * 100:.1f}%"),
            ("", ""),
            ("Beta", f"{results.beta:.2f}"),
            ("Alpha", f"{results.alpha * 100:.1f}%"),
            ("", ""),
            ("Total Trades", results.total_trades),
            ("Winning Trades", results.winning_trades),
            ("Losing Trades", results.losing_trades),
            ("Win Rate", f"{results.win_rate * 100:.1f}%"),
            ("Transaction Costs", f"${results.total_transaction_costs:,.2f}"),
        ]

        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)

        # Portfolio Values Sheet
        ws_values = wb.create_sheet("Portfolio Values")
        if len(results.portfolio_values) > 0:
            ws_values.append(["Date", "Portfolio Value", "Cash", "Holdings Value"])
            for idx, row in results.portfolio_values.iterrows():
                ws_values.append([
                    idx.strftime("%Y-%m-%d"),
                    row["portfolio_value"],
                    row["cash"],
                    row["holdings_value"]
                ])

        # Trade History Sheet
        ws_trades = wb.create_sheet("Trade History")
        ws_trades.append(["Date", "Ticker", "Action", "Shares", "Price", "Value", "Cost", "Reason"])
        for trade in results.trades:
            ws_trades.append([
                trade.date.strftime("%Y-%m-%d"),
                trade.ticker,
                trade.action,
                trade.shares,
                f"${trade.price:.2f}",
                f"${trade.value:,.2f}",
                f"${trade.transaction_cost:.2f}",
                trade.reason
            ])

        wb.save(filepath)
        return filepath
